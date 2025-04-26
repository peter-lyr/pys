try:
    import pandas as pd
    import re
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill
except:
    import os
    os.system('pip install pandas openpyxl -i https://pypi.tuna.tsinghua.edu.cn/simple --trusted-host mirrors.aliyun.com')
    import pandas as pd
    import re
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill


def markdown_to_excel(markdown_text):
    rows = []
    current_level = 0
    current_content = ""
    current_title = ""
    title_stack = []

    lines = markdown_text.split('\n')
    for i, line in enumerate(lines):
        line = line.strip()
        if re.match(r'^[=-]{3,}$', line):
            # 检查上一行是否为标题
            if i > 0:
                prev_line = lines[i - 1].strip()
                if prev_line:
                    if current_content:
                        rows.append([*title_stack, current_content])
                        current_content = ""
                    current_title = prev_line
                    if line.startswith('='):
                        current_level = 1
                        title_stack = [current_title]
                    else:
                        current_level = 2
                        if len(title_stack) > 0:
                            title_stack = title_stack[:1] + [current_title]
                        else:
                            title_stack = [current_title]
        elif line.startswith("#"):
            if current_content:
                rows.append([*title_stack, current_content])
                current_content = ""
            level = line.count("#")
            current_title = line[level:].strip()
            current_level = level
            if len(title_stack) >= level:
                title_stack = title_stack[:level - 1] + [current_title]
            else:
                title_stack.extend([current_title] * (level - len(title_stack)))
        else:
            current_content += "\n" + line if current_content else line

    if current_content:
        rows.append([*title_stack, current_content])

    # 处理不同层级标题列数不一致的情况
    max_columns = max(len(row) for row in rows)
    for row in rows:
        if len(row) < max_columns:
            row.extend([''] * (max_columns - len(row)))

    df = pd.DataFrame(rows)
    print("处理后数据 DataFrame:")
    print(df)

    return df


def merge_cells(workbook, sheet_name):
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    max_col = sheet.max_column

    for col in range(1, max_col + 1):
        start_row = 1
        for row in range(2, max_row + 1):
            if sheet.cell(row, col).value == sheet.cell(start_row, col).value:
                continue
            else:
                if row - start_row > 1:
                    sheet.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col)}{row - 1}')
                    sheet.cell(start_row, col).alignment = Alignment(vertical='center')
                start_row = row
        if max_row - start_row > 0:
            sheet.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col)}{max_row}')
            sheet.cell(start_row, col).alignment = Alignment(vertical='center')


def set_paragraph_background(workbook, sheet_name, df):
    sheet = workbook[sheet_name]
    for row_idx, row in df.iterrows():
        # 从最后一列开始往前找
        for col_idx in range(df.shape[1], 0, -1):
            cell_value = row[col_idx - 1]
            if cell_value:
                cell = sheet.cell(row=row_idx + 1, column=col_idx)
                fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                cell.fill = fill
                break


try:
    with open(r'C:\Users\llydr\w\work_summary_week.md', 'r', encoding='utf-8') as file:
        markdown_text = file.read()
    result_df = markdown_to_excel(markdown_text)

    # 创建一个新的 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    # 将 DataFrame 数据写入工作表
    for r_idx, row in enumerate(result_df.values, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 合并相同内容的单元格
    merge_cells(wb, ws.title)

    # 设置普通段落背景为黄色
    set_paragraph_background(wb, ws.title, result_df)

    # 保存工作簿
    wb.save(r'C:\Users\llydr\w\work_summary_week.md.xlsx')
    print(r"转换成功，已保存为 C:\Users\llydr\w\work_summary_week.md.xlsx")
except FileNotFoundError:
    print("未找到指定的 Markdown 文件，请检查文件路径。")
except Exception as e:
    print(f"发生错误: {e}")
