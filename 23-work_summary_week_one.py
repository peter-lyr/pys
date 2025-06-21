import os
import re
import sys

try:
    import pandas as pd
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except:
    os.system(
        "pip install pandas openpyxl -i https://pypi.tuna.tsinghua.edu.cn/simple --trusted-host mirrors.aliyun.com"
    )
    import pandas as pd
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter


def markdown_to_excel(markdown_text):
    rows = []
    current_title = ""
    title_stack = []
    lines = markdown_text.split("\n")
    prev_line_is_title = False
    current_paragraph = ""
    for i, line in enumerate(lines):
        line = line.strip()
        if re.match(r"^[=-]{3,}$", line):
            if prev_line_is_title:
                # 处理标题分隔符，前提是上一行是标题
                if i > 0:
                    prev_line = lines[i - 1].strip()
                    if prev_line:
                        if current_paragraph:
                            rows.append([*title_stack, current_paragraph])
                            current_paragraph = ""
                        current_title = prev_line
                        if line.startswith("="):
                            title_stack = [current_title]
                        else:
                            if len(title_stack) > 0:
                                title_stack = title_stack[:1] + [current_title]
                            else:
                                title_stack = [current_title]
                prev_line_is_title = False
            else:
                # 若上一行不是标题，将分隔符作为普通内容处理
                if current_paragraph:
                    rows.append([*title_stack, current_paragraph])
                    current_paragraph = ""
                rows.append([*title_stack, line])
                prev_line_is_title = False
        elif line.startswith("#"):
            # 处理 Markdown 标题
            if current_paragraph:
                rows.append([*title_stack, current_paragraph])
                current_paragraph = ""
            level = line.count("#")
            current_title = line[level:].strip()
            if len(title_stack) >= level:
                title_stack = title_stack[: level - 1] + [current_title]
            else:
                title_stack.extend([current_title] * (level - len(title_stack)))
            prev_line_is_title = True
        elif line:
            # 处理普通段落内容
            if current_paragraph:
                current_paragraph += "\n" + line
            else:
                current_paragraph = line
            prev_line_is_title = False
        else:
            # 遇到空行，将当前段落添加到结果中
            if current_paragraph:
                rows.append([*title_stack, current_paragraph])
                current_paragraph = ""
            prev_line_is_title = False

    # 处理最后一个段落
    if current_paragraph:
        rows.append([*title_stack, current_paragraph])

    # 处理不同层级标题列数不一致的情况
    max_columns = max(len(row) for row in rows)
    for row in rows:
        if len(row) < max_columns:
            row.extend([None] * (max_columns - len(row)))

    # 增加段落行数记录列
    for i, row in enumerate(rows):
        # 找到最后一个有值的列作为普通段落列
        for j in range(len(row) - 1, -1, -1):
            if row[j]:
                line_count = len(row[j].split("\n"))
                rows[i] = row[:j] + [line_count] + row[j:]
                break

    df = pd.DataFrame(rows)
    # print("处理后数据 DataFrame:")
    # print(df)
    return df


def set_cell_styles(workbook, sheet_name):
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    max_col = sheet.max_column

    # 定义不同级别的标题样式，使用更易辨认的颜色和调整后的字体大小，字体设置为 Arial Black
    title_styles = {
        1: {
            "font": Font(name="Arial Black", bold=True, size=13, color="FFFFFF"),
            "fill": PatternFill(
                start_color="FF6666", end_color="FF6666", fill_type="solid"
            ),
            "alignment": Alignment(vertical="top", horizontal="left"),
        },
        2: {
            "font": Font(name="Arial Black", bold=True, size=12, color="FFFFFF"),
            "fill": PatternFill(
                start_color="FF9933", end_color="FF9933", fill_type="solid"
            ),
            "alignment": Alignment(vertical="top", horizontal="left"),
        },
        3: {
            "font": Font(name="Arial Black", bold=True, size=11, color="FFFFFF"),
            "fill": PatternFill(
                start_color="66CC66", end_color="66CC66", fill_type="solid"
            ),
            "alignment": Alignment(vertical="top", horizontal="left"),
        },
        4: {
            "font": Font(name="Arial Black", bold=True, size=10, color="000000"),
            "fill": PatternFill(
                start_color="99CCFF", end_color="99CCFF", fill_type="solid"
            ),
            "alignment": Alignment(vertical="top", horizontal="left"),
        },
        5: {
            "font": Font(name="Arial Black", bold=True, size=10, color="FFFFFF"),
            "fill": PatternFill(
                start_color="CC99FF", end_color="CC99FF", fill_type="solid"
            ),
            "alignment": Alignment(vertical="top", horizontal="left"),
        },
        6: {
            "font": Font(name="Arial Black", bold=True, size=10, color="000000"),
            "fill": PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            ),
            "alignment": Alignment(vertical="top", horizontal="left"),
        },
    }

    # 定义普通段落样式，字体设置为 Arial Black
    paragraph_style = {
        "font": Font(name="Arial Black", size=10, color="000000"),
        "fill": PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        ),
        "alignment": Alignment(vertical="top", horizontal="left"),
    }

    # 定义段落行数记录列样式，字体设置为 Arial Black
    line_count_style = {
        "font": Font(name="Arial Black", bold=True, size=10, color="000000"),
        "fill": PatternFill(
            start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
        ),
        "alignment": Alignment(vertical="top", horizontal="center"),
    }

    # 定义边框样式，将边框设置为 thin 使其更明显
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in range(1, max_row + 1):
        last_col = 0
        for col in range(1, max_col + 1):
            cell = sheet.cell(row, col)
            last_col = col - 1
            if not cell.value or pd.isnull(cell.value):
                break
        else:
            last_col = max_col
        for col in range(1, max_col + 1):
            cell = sheet.cell(row, col)
            if cell.value:
                if col < last_col - 1:
                    # 设置标题样式
                    cell.font = title_styles[col]["font"]
                    cell.fill = title_styles[col]["fill"]
                    cell.alignment = title_styles[col]["alignment"]
                elif col == last_col - 1:
                    # 段落行数记录列样式
                    cell.font = line_count_style["font"]
                    cell.fill = line_count_style["fill"]
                    cell.alignment = line_count_style["alignment"]
                else:
                    # 设置普通段落样式
                    cell.font = paragraph_style["font"]
                    cell.fill = paragraph_style["fill"]
                    cell.alignment = paragraph_style["alignment"]
                # 在合并单元格之前设置边框
                cell.border = border


def merge_cells(workbook, sheet_name):
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    max_col = sheet.max_column
    for col in range(1, max_col - 1):
        start_row = 1
        for row in range(2, max_row + 1):
            if (
                sheet.cell(row, col).value
                and sheet.cell(row, col).value == sheet.cell(start_row, col).value
            ):
                continue
            else:
                if row - start_row > 1 and sheet.cell(start_row, col).value:
                    sheet.merge_cells(
                        f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{row - 1}"
                    )
                    top_left_cell = sheet.cell(start_row, col)
                    top_left_cell.alignment = Alignment(vertical="center")
                start_row = row
        if max_row - start_row > 0 and sheet.cell(start_row, col).value:
            sheet.merge_cells(
                f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{max_row}"
            )
            top_left_cell = sheet.cell(start_row, col)
            top_left_cell.alignment = Alignment(vertical="center")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("请提供 Markdown 文件的路径作为参数。")
        sys.exit(1)

    markdown_file_path = sys.argv[1]
    week_range = sys.argv[2]
    week_num = week_range.split(" ")[0]
    print(week_num)
    try:
        with open(markdown_file_path, "r", encoding="utf-8") as file:
            markdown_text = file.read()
        result_df = markdown_to_excel(markdown_text)

        D = {}
        for r_idx, row in enumerate(result_df.values, start=1):
            for c_idx, value in enumerate(row, start=1):
                print(value)
                if not value:
                    continue
                if c_idx == 2:
                    # print(f'{row}')
                    # print('->'.join([str(i) for i in row]))
                    # print(len(row))
                    if (type(row[2]) == str):
                        # print(f"{str(row[0]):10s} | {str(row[1]):60s} | {str(row[2]):60s} | {str(row[4]):60s}")
                        # print(f"{str(row[1])} | {str(row[2])} | {str(row[4])}")
                        # print()
                        task = row[1]
                        date = row[2]
                        detail = row[4]
                        main_task = task.split(' ')[0]
                        sub_task = ' '.join(task.split(' ')[1:])
                        if main_task not in D:
                            D[main_task] = {}
                        S = {}
                        if sub_task:
                            if sub_task not in D[main_task]:
                                D[main_task][sub_task] = {}
                            S = D[main_task][sub_task]
                        else:
                            S = D[main_task]
                        S[date] = detail
        # print(D)
        # print()
        with open(f'{os.path.splitext(markdown_file_path)[0]}_one.md', 'wb') as f:
            K1 = D.keys()
            K1 = sorted(K1)
            # print(f'[toc]')
            f.write(f'[toc]\n'.encode('utf-8'))
            for k1 in K1:
                v1 = D[k1]
                # print(f'\n# {k1}')
                f.write(f'\n# {k1}\n'.encode('utf-8'))
                K2 = v1.keys()
                K2 = sorted(K2, reverse=True)
                for k2 in K2:
                    v2 = v1[k2]
                    if type(v2) is dict:
                        # print(f'## {k2}')
                        f.write(f'## {k2}\n'.encode('utf-8'))
                        K3 = v2.keys()
                        K3 = sorted(K3)
                        for k3 in K3:
                            v3 = v2[k3]
                            # print(f'#### {k3}')
                            f.write(f'#### {k3}\n'.encode('utf-8'))
                            # print(f'{v3}')
                            f.write(f'{v3}\n'.encode('utf-8'))
                    else:
                        # print(f'### {k2}')
                        f.write(f'### {k2}\n'.encode('utf-8'))
                        # print(f'{v2}')
                        f.write(f'{v2}\n'.encode('utf-8'))
    except FileNotFoundError:
        print("未找到指定的 Markdown 文件，请检查文件路径。")
    except Exception as e:
        print(f"发生错误: {e}")
