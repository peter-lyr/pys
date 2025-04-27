import sys
import os

try:
    import pandas as pd
    import re
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill
except:
    import os
    os.system('pip install pandas openpyxl -i https://pypi.tuna.tsinghua.edu.cn/simple --trusted-host mirrors.aliyun.com')
    import pandas as pd
    import re
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, PatternFill


def markdown_to_excel(markdown_text):
    rows = []
    current_level = 0
    current_title = ""
    title_stack = []
    lines = markdown_text.split('\n')
    prev_line_is_title = False
    for i, line in enumerate(lines):
        line = line.strip()
        if re.match(r'^[=-]{3,}$', line):
            if prev_line_is_title:
                if i > 0:
                    prev_line = lines[i - 1].strip()
                    if prev_line:
                        current_title = prev_line
                        if line.startswith('='):
                            current_level = 1
                            title_stack = [current_title]
                        else:
                            current_level = 2
                            if len(title_stack) > 0:
                                title_stack = title_stack[:1] + [current_title]
                            else:
                                title_stack = [current_title]
                prev_line_is_title = False
            else:
                rows.append([*title_stack, line])
                prev_line_is_title = False
        elif line.startswith("#"):
            level = line.count("#")
            current_title = line[level:].strip()
            current_level = level
            if len(title_stack) >= level:
                title_stack = title_stack[:level - 1] + [current_title]
            else:
                title_stack.extend([current_title] * (level - len(title_stack)))
            prev_line_is_title = True
        elif line:
            paragraphs = line.split('\n\n')
            for paragraph in paragraphs:
                if paragraph:
                    rows.append([*title_stack, paragraph])
            prev_line_is_title = False
        else:
            prev_line_is_title = False

    max_columns = max(len(row) for row in rows)
    for row in rows:
        if len(row) < max_columns:
            row.extend([None] * (max_columns - len(row)))

    df = pd.DataFrame(rows)
    print("处理后数据 DataFrame:")
    print(df)
    return df


def merge_cells(workbook, sheet_name):
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    max_col = sheet.max_column
    for col in range(1, max_col + 1):
        start_row = 1
        for row in range(2, max_row + 1):
            if sheet.cell(row, col).value and sheet.cell(row, col).value == sheet.cell(start_row, col).value:
                continue
            else:
                if row - start_row > 1 and sheet.cell(start_row, col).value:
                    sheet.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col)}{row - 1}')
                    sheet.cell(start_row, col).alignment = Alignment(vertical='center')
                start_row = row
        if max_row - start_row > 0 and sheet.cell(start_row, col).value:
            sheet.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col)}{max_row}')
            sheet.cell(start_row, col).alignment = Alignment(vertical='center')


def set_paragraph_background(workbook, sheet_name, df):
    sheet = workbook[sheet_name]
    for row_idx, row in df.iterrows():
        for col_idx in range(df.shape[1], 0, -1):
            cell_value = row[col_idx - 1]
            if cell_value:
                cell = sheet.cell(row=row_idx + 1, column=col_idx)
                fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                cell.fill = fill
                break


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("请提供 Markdown 文件的路径作为参数。")
        sys.exit(1)

    markdown_file_path = sys.argv[1]
    try:
        with open(markdown_file_path, 'r', encoding='utf-8') as file:
            markdown_text = file.read()
        result_df = markdown_to_excel(markdown_text)

        wb = Workbook()
        ws = wb.active
        for r_idx, row in enumerate(result_df.values, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        merge_cells(wb, ws.title)
        set_paragraph_background(wb, ws.title, result_df)

        file_dir = os.path.dirname(markdown_file_path)
        file_name = os.path.basename(markdown_file_path)
        excel_file_name = os.path.splitext(file_name)[0] + ".xlsx"
        excel_file_path = os.path.join(file_dir, excel_file_name)

        wb.save(excel_file_path)
        print(f"转换成功，已保存为 {excel_file_path}")
    except FileNotFoundError:
        print("未找到指定的 Markdown 文件，请检查文件路径。")
    except Exception as e:
        print(f"发生错误: {e}")
